"""One-time script: download EPA eGRID 2022 and build data/carbon/egrid_2022.json
and data/carbon/egrid_2022_subregions.json.

Usage:
    python scripts/build_egrid_json.py

Requires: openpyxl (dev extra), requests (runtime dep).
Network call: downloads eGRID2022_data.xlsx from EPA once; result committed to the repo.
"""
from __future__ import annotations

import hashlib
import json
import sys
from pathlib import Path

import requests

# ── Source (P2) ───────────────────────────────────────────────────────────────
_URL = (
    "https://www.epa.gov/system/files/documents/2024-01/egrid2022_data.xlsx"
)
_RETRIEVAL_DATE = "2026-06-10"

# State-level sheet
_SHEET = "ST22"
_STATE_COL = "PSTATABB"  # 2-letter state abbreviation
_FACTOR_COL = "STC2ERTA" # state annual CO2e total output emission rate, lb/MWh

# Subregion sheet (R6-2 / B2)
_SR_SHEET = "SRL22"
_SR_ACRONYM_COL = "SUBRGN"   # subregion acronym
_SR_FACTOR_COL = "SRC2ERTA"  # subregion annual CO2e total output emission rate, lb/MWh

# US territories (PR, VI, ...) included in the state sheet but outside OpenUBEM's continental scope
_EXCLUDE_STATES = frozenset({"PR", "VI", "AS", "GU", "MP"})

# City → subregion mapping for the 3 R6-2 cities (informational tag in state JSON)
_CITY_SUBREGIONS = {"NY": "NYCW", "CA": "CAMX", "TX": "ERCT"}

# lb/MWh → kg/kWh: ×0.453592 (lb→kg) ÷ 1000 (MWh→kWh)
_LB_MWH_TO_KG_KWH = 0.453592 / 1000.0

_OUT = Path(__file__).parent.parent / "openubem" / "data" / "carbon" / "egrid_2022.json"
_OUT_SUBREGIONS = Path(__file__).parent.parent / "openubem" / "data" / "carbon" / "egrid_2022_subregions.json"
_PROV = Path(__file__).parent.parent / "openubem" / "data" / "carbon" / "PROVENANCE.md"


def _download(url: str) -> bytes:
    print(f"Downloading {url} ...", flush=True)
    resp = requests.get(url, timeout=120)
    resp.raise_for_status()
    return resp.content


def _build_table(xlsx_bytes: bytes) -> dict[str, dict]:
    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[_SHEET]

    rows = list(ws.iter_rows(values_only=True))
    # Row 0 = long-form headers, row 1 = short-code headers (PSTATABB etc.), row 2+ = data
    shortcode_header = [str(c).strip() if c is not None else "" for c in rows[1]]

    try:
        i_state = shortcode_header.index(_STATE_COL)
        i_factor = shortcode_header.index(_FACTOR_COL)
    except ValueError as exc:
        raise RuntimeError(
            f"Expected short-code columns {_STATE_COL!r}, {_FACTOR_COL!r} "
            f"in row 1 of sheet {_SHEET!r}. Found: {shortcode_header}"
        ) from exc

    result: dict[str, dict] = {}
    for row in rows[2:]:
        state = row[i_state]
        factor_raw = row[i_factor]
        if state is None or factor_raw is None:
            continue
        state = str(state).strip().upper()
        if len(state) != 2 or state in _EXCLUDE_STATES:
            continue
        try:
            factor_lb_mwh = float(factor_raw)
        except (TypeError, ValueError):
            continue
        factor_kg_kwh = round(factor_lb_mwh * _LB_MWH_TO_KG_KWH, 6)
        result[state] = {
            "subregion": _CITY_SUBREGIONS.get(state, ""),  # populated for NY/CA/TX (B2 decision)
            "factor_kgco2_kwh": factor_kg_kwh,
        }

    wb.close()
    return result


def _build_subregion_table(xlsx_bytes: bytes) -> dict[str, dict]:
    """Extract subregion CO2e total-output factors from the SRL22 sheet (R6-2 / B2)."""
    import openpyxl
    import io

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    if _SR_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"Expected subregion sheet {_SR_SHEET!r} not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[_SR_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    shortcode_header = [str(c).strip() if c is not None else "" for c in rows[1]]

    try:
        i_acronym = shortcode_header.index(_SR_ACRONYM_COL)
        i_factor = shortcode_header.index(_SR_FACTOR_COL)
    except ValueError as exc:
        raise RuntimeError(
            f"Expected short-code columns {_SR_ACRONYM_COL!r}, {_SR_FACTOR_COL!r} "
            f"in row 1 of sheet {_SR_SHEET!r}. Found: {shortcode_header}"
        ) from exc

    result: dict[str, dict] = {}
    for row in rows[2:]:
        acronym = row[i_acronym]
        factor_raw = row[i_factor]
        if acronym is None or factor_raw is None:
            continue
        acronym = str(acronym).strip()
        if not acronym:
            continue
        try:
            factor_lb_mwh = float(factor_raw)
        except (TypeError, ValueError):
            continue
        factor_kg_kwh = round(factor_lb_mwh * _LB_MWH_TO_KG_KWH, 6)
        result[acronym] = {"factor_kgco2_kwh": factor_kg_kwh}

    wb.close()
    return result


def _validate(table: dict) -> None:
    assert len(table) == 51, f"Expected 51 entries (50 states + DC), got {len(table)}"
    assert "MA" in table, "Massachusetts (MA) missing"
    for state, rec in table.items():
        f = rec["factor_kgco2_kwh"]
        assert 0.01 < f < 1.2, (
            f"{state}: factor {f} outside (0.01, 1.2) kg/kWh plausibility band"
        )
    # Verify NY/CA/TX subregion tags populated (B2 decision)
    for st, expected_sr in _CITY_SUBREGIONS.items():
        assert table[st]["subregion"] == expected_sr, (
            f"{st}: expected subregion={expected_sr!r}, got {table[st]['subregion']!r}"
        )
    print(f"Validation OK: {len(table)} entries, MA factor = {table['MA']['factor_kgco2_kwh']}")


def _validate_subregions(table: dict) -> None:
    required = {"NYCW", "CAMX", "ERCT"}
    for acronym in required:
        assert acronym in table, f"Required subregion {acronym!r} missing from {_SR_SHEET}"
        f = table[acronym]["factor_kgco2_kwh"]
        assert 0.01 < f < 1.2, (
            f"{acronym}: factor {f} outside (0.01, 1.2) kg/kWh plausibility band"
        )
    print(
        f"Subregion validation OK: {len(table)} subregions; "
        f"NYCW={table['NYCW']['factor_kgco2_kwh']}, "
        f"CAMX={table['CAMX']['factor_kgco2_kwh']}, "
        f"ERCT={table['ERCT']['factor_kgco2_kwh']}"
    )


def main() -> None:
    _OUT.parent.mkdir(parents=True, exist_ok=True)

    xlsx_bytes = _download(_URL)
    sha256 = hashlib.sha256(xlsx_bytes).hexdigest()
    print(f"SHA-256: {sha256}", flush=True)

    # State-level table
    table = _build_table(xlsx_bytes)
    _validate(table)

    _OUT.write_text(json.dumps(table, indent=2, sort_keys=True), encoding="utf-8")
    print(f"Written: {_OUT} ({len(table)} entries)", flush=True)

    # Subregion table (R6-2 / B2)
    sr_table = _build_subregion_table(xlsx_bytes)
    _validate_subregions(sr_table)

    _OUT_SUBREGIONS.write_text(json.dumps(sr_table, indent=2, sort_keys=True), encoding="utf-8")
    print(f"Written: {_OUT_SUBREGIONS} ({len(sr_table)} subregions)", flush=True)

    prov = f"""# PROVENANCE — openubem/data/carbon/egrid_2022.json + egrid_2022_subregions.json

## Source
- Dataset: EPA eGRID 2022 (Emissions & Generation Resource Integrated Database)
- URL: {_URL}
- Retrieval date: {_RETRIEVAL_DATE}
- SHA-256 of downloaded xlsx: {sha256}

## State-level factors (egrid_2022.json)
- Excel sheet: {_SHEET} (state-level summary)
- Short-code header row (row 1): {_STATE_COL} = 2-letter USPS abbreviation; {_FACTOR_COL} = state annual CO₂e total output emission rate (lb/MWh)
- Territories excluded: {sorted(_EXCLUDE_STATES)} (outside OpenUBEM continental-US scope)
- factor_kgco2_kwh = {_FACTOR_COL}_lb_mwh × (0.453592 lb/kg) ÷ 1000 (MWh/kWh)
- 51 entries: 50 US states + DC
- MA factor: {table.get('MA', {}).get('factor_kgco2_kwh', 'N/A')} kg CO₂e/kWh
- NY subregion tag: {table.get('NY', {}).get('subregion', '')} (informational, B2 decision 2026-06-15)
- CA subregion tag: {table.get('CA', {}).get('subregion', '')} (informational, B2 decision 2026-06-15)
- TX subregion tag: {table.get('TX', {}).get('subregion', '')} (informational, B2 decision 2026-06-15)

## Subregion factors (egrid_2022_subregions.json) — R6-2 / B2 decision 2026-06-15
- Excel sheet: {_SR_SHEET} (subregion summary)
- Short-code header row (row 1): {_SR_ACRONYM_COL} = eGRID subregion acronym; {_SR_FACTOR_COL} = subregion annual CO₂e total output emission rate (lb/MWh)
- factor_kgco2_kwh = {_SR_FACTOR_COL}_lb_mwh × (0.453592 lb/kg) ÷ 1000 (MWh/kWh)
- {len(sr_table)} subregions extracted
- City→subregion mapping for R6-2: NYC→NYCW ({sr_table.get('NYCW', {}).get('factor_kgco2_kwh', 'N/A')} kg CO₂e/kWh), LA→CAMX ({sr_table.get('CAMX', {}).get('factor_kgco2_kwh', 'N/A')} kg CO₂e/kWh), Austin→ERCT ({sr_table.get('ERCT', {}).get('factor_kgco2_kwh', 'N/A')} kg CO₂e/kWh)

## B2 decision (PLAN R6 §4.3, manager ruling 2026-06-15)
GWP recompute uses grid-subregion factors (from egrid_2022_subregions.json) instead of
state-level factors for the 3 R6 cities. Heating GWP (natural gas) is unchanged. Only
cooling + lighting + equipment electricity GWP is rescaled by ratio = f_subregion / f_state.
This is post-processing of energy already in 05_results.csv — no resimulation.
R5 shipped state-level GWP remains the immutable baseline; subregion GWP presented as R6 refinement.

## Downstream use (DESIGN §3E, PLAN F7)
- Heating GWP: × 0.181 kg CO₂e/kWh (natural gas, Iseri et al. 2025)
- Cooling / Lighting / Equipment GWP: × egrid_2022[state]['factor_kgco2_kwh'] (runtime core)
- R6-2 post-processing only: × egrid_2022_subregions[acronym]['factor_kgco2_kwh'] applied in r6_rescore_cells.py
- Convention: load_referenced_v1 (no η or COP applied — see DESIGN §3E)
"""
    _PROV.write_text(prov, encoding="utf-8")
    print(f"Written: {_PROV}", flush=True)

    print("\nMA entry:", json.dumps(table["MA"], indent=2))
    print("\nNY entry:", json.dumps(table["NY"], indent=2))
    print("\nKey subregion entries:")
    for a in ("NYCW", "CAMX", "ERCT"):
        print(f"  {a}:", json.dumps(sr_table[a], indent=2))


if __name__ == "__main__":
    main()
